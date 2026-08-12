#!/usr/bin/env python3
"""Утилита замены employeeNumber в JSON-выгрузках по CSV-справочнику."""

from __future__ import annotations

import argparse
import csv
import json
import random
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "[ERROR] Нужен пакет openpyxl. Установите: pip install openpyxl\n"
        "или: pip install -r requirements.txt"
    ) from exc


@dataclass(frozen=True)
class EmployeeRow:
    person_number_8: str
    surname: str
    first_name: str
    gosb_code: str
    tb_code: str


@dataclass(frozen=True)
class ReplacementOccurrence:
    """Одно вхождение исходного сотрудника в файле/новости (снимок «было»)."""

    source_emp: str
    file_name: str
    news_id: str
    last_name: str
    first_name: str
    gosb_code: str
    tb_code: str
    ter_division_name: str


@dataclass(frozen=True)
class CsvFilterRule:
    """Одно правило фильтрации CSV-строк."""

    business_block: str
    role_code: str
    uch_code: str | None


# Порядок полей в Excel: сначала все «было», затем все «стало», затем новости, затем файлы.
BEFORE_FIELD_KEYS: tuple[str, ...] = (
    "employeeNumber",
    "lastName",
    "firstName",
    "gosbCode",
    "tbCode",
    "terDivisionName",
)
EXCEL_HEADERS: tuple[str, ...] = (
    "Было employeeNumber",
    "Было lastName",
    "Было firstName",
    "Было gosbCode",
    "Было tbCode",
    "Было terDivisionName",
    "Стало employeeNumber",
    "Стало lastName",
    "Стало firstName",
    "Стало gosbCode",
    "Стало tbCode",
    "Стало terDivisionName",
    "ID новостей",
    "Файлы",
)


def normalize_person_number(raw: Any) -> str:
    """Привести табельный к формату из 8 цифр."""
    digits: str = "".join(ch for ch in str(raw or "") if ch.isdigit())
    if not digits:
        return ""
    return digits[-8:].zfill(8)


def load_config(config_path: Path) -> dict[str, Any]:
    """Загрузить конфиг из JSON-файла."""
    with config_path.open("r", encoding="utf-8") as fp:
        cfg: dict[str, Any] = json.load(fp)
    return cfg


def resolve_config_path(cli_config: str | None) -> Path:
    """
    Найти путь к конфигу.
    Приоритет:
    1) явный --config;
    2) config_emp_replace.json рядом со скриптом;
    3) первый config_*.json рядом со скриптом.
    """
    if cli_config:
        path = Path(cli_config).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"Конфиг не найден: {path}")
        return path

    script_dir: Path = Path(__file__).resolve().parent
    preferred: Path = script_dir / "config_emp_replace.json"
    if preferred.is_file():
        return preferred

    candidates: list[Path] = sorted(script_dir.glob("config_*.json"))
    if candidates:
        return candidates[0]

    raise FileNotFoundError(
        "Конфиг не указан и не найден рядом со скриптом "
        "(ожидается config_emp_replace.json или config_*.json). "
        "Можно передать путь явно: --config путь/к/config.json"
    )


def load_employee_pool(csv_path: Path, cfg: dict[str, Any]) -> dict[str, EmployeeRow]:
    """Считать пул табельных из CSV с фильтрами."""
    delimiter: str = str(cfg.get("csv_delimiter", ";"))
    filter_rules: list[CsvFilterRule] = build_csv_filter_rules(cfg)
    pool: dict[str, EmployeeRow] = {}
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp, delimiter=delimiter)
        for row in reader:
            if not any(is_row_matches_rule(row, rule) for rule in filter_rules):
                continue
            person_number_8: str = normalize_person_number(row.get("PERSON_NUMBER", ""))
            if not person_number_8:
                continue
            pool[person_number_8] = EmployeeRow(
                person_number_8=person_number_8,
                surname=str(row.get("SURNAME", "")).strip(),
                first_name=str(row.get("FIRST_NAME", "")).strip(),
                gosb_code=str(row.get("GOSB_CODE", "")).strip(),
                tb_code=str(row.get("TB_CODE", "")).strip(),
            )
    return pool


def normalize_filter_values(raw_value: Any) -> list[str]:
    """Нормализовать значение фильтра к списку непустых строк."""
    if raw_value is None:
        return []
    if isinstance(raw_value, list):
        values: list[str] = []
        for item in raw_value:
            text: str = str(item or "").strip()
            if text:
                values.append(text)
        return values
    text = str(raw_value).strip()
    if not text:
        return []
    return [text]


def is_any_uch_value(raw_uch_code: str) -> bool:
    """Проверить, что в правиле указан «любой UCH_CODE»."""
    return raw_uch_code.strip().upper() in {"*", "ANY", "ALL"}


def build_csv_filter_rules(cfg: dict[str, Any]) -> list[CsvFilterRule]:
    """
    Построить правила фильтрации CSV.
    Поддержка:
    - filter_rules: [{business_block, role_code, uch_code}]
    - legacy-поля filter_business_block/filter_role_code/filter_uch_code (строка или список).
    """
    raw_rules: Any = cfg.get("filter_rules")
    parsed_rules: list[CsvFilterRule] = []
    if isinstance(raw_rules, list):
        for item in raw_rules:
            if not isinstance(item, dict):
                continue
            business_block: str = str(item.get("business_block", "")).strip()
            role_code: str = str(item.get("role_code", "")).strip()
            raw_uch_code: str = str(item.get("uch_code", "")).strip()
            if not business_block or not role_code:
                continue
            uch_code: str | None = None if (not raw_uch_code or is_any_uch_value(raw_uch_code)) else raw_uch_code
            parsed_rules.append(
                CsvFilterRule(
                    business_block=business_block,
                    role_code=role_code,
                    uch_code=uch_code,
                )
            )
    if parsed_rules:
        return parsed_rules

    business_blocks: list[str] = normalize_filter_values(
        cfg.get("filter_business_block", "KMKKSB")
    )
    role_codes: list[str] = normalize_filter_values(cfg.get("filter_role_code", "KM_KKSB"))
    uch_codes_raw: list[str] = normalize_filter_values(cfg.get("filter_uch_code", "1"))

    if not business_blocks:
        business_blocks = ["KMKKSB"]
    if not role_codes:
        role_codes = ["KM_KKSB"]
    if not uch_codes_raw:
        uch_codes_raw = ["1"]

    uch_codes: list[str | None] = []
    for raw_uch_code in uch_codes_raw:
        if is_any_uch_value(raw_uch_code):
            uch_codes.append(None)
        else:
            uch_codes.append(raw_uch_code)
    if not uch_codes:
        uch_codes = ["1"]

    for business_block in business_blocks:
        for role_code in role_codes:
            for uch_code in uch_codes:
                parsed_rules.append(
                    CsvFilterRule(
                        business_block=business_block,
                        role_code=role_code,
                        uch_code=uch_code,
                    )
                )
    return parsed_rules


def is_row_matches_rule(row: dict[str, Any], rule: CsvFilterRule) -> bool:
    """Проверить соответствие CSV-строки одному правилу фильтра."""
    if str(row.get("BUSINESS_BLOCK", "")).strip() != rule.business_block:
        return False
    if str(row.get("ROLE_CODE", "")).strip() != rule.role_code:
        return False
    if rule.uch_code is None:
        return True
    return str(row.get("UCH_CODE", "")).strip() == rule.uch_code


def extract_news_id(node: dict[str, Any]) -> str:
    """
    Извлечь ID новости из объекта, похожего на news-item.
    objectId / newsId при наличии типичных полей новости.
    """
    news_markers: tuple[str, ...] = (
        "leadersList",
        "authorsList",
        "newsText",
        "newsStatus",
        "summary",
        "description",
        "newsType",
        "newsTagList",
        "businessBlock",
    )
    if not any(marker in node for marker in news_markers):
        return ""
    for key in ("objectId", "newsId", "id"):
        raw: Any = node.get(key)
        if raw is None:
            continue
        text: str = str(raw).strip()
        if text:
            return text
    return ""


def iter_employee_occurrences(
    node: Any, news_id: str = ""
) -> Iterator[tuple[dict[str, Any], str]]:
    """Обойти дерево: каждый узел с employeeNumber + текущий ID новости."""
    if isinstance(node, dict):
        local_news_id: str = extract_news_id(node) or news_id
        if "employeeNumber" in node:
            yield node, local_news_id
        for value in node.values():
            yield from iter_employee_occurrences(value, local_news_id)
    elif isinstance(node, list):
        for item in node:
            yield from iter_employee_occurrences(item, news_id)


def walk_employee_nodes(node: Any) -> list[dict[str, Any]]:
    """Найти все объекты, где присутствует employeeNumber."""
    return [emp_node for emp_node, _news_id in iter_employee_occurrences(node)]


def collect_input_files(input_dir: Path, cfg: dict[str, Any]) -> list[Path]:
    """
    Собрать входные JSON.
    Приоритет: cfg.input_files (список имён) → иначе input_glob в input_dir.
    """
    raw_names: Any = cfg.get("input_files")
    files: list[Path] = []
    seen: set[str] = set()

    if isinstance(raw_names, list) and raw_names:
        for item in raw_names:
            name: str = str(item or "").strip()
            if not name:
                continue
            path = Path(name)
            if not path.is_absolute():
                path = input_dir / path
            path = path.resolve()
            key: str = str(path)
            if key in seen:
                continue
            seen.add(key)
            if not path.is_file():
                raise FileNotFoundError(f"Входной файл не найден: {path}")
            files.append(path)
        return files

    pattern: str = str(cfg.get("input_glob", "*.json"))
    for path in sorted(input_dir.glob(pattern)):
        if not path.is_file():
            continue
        key = str(path.resolve())
        if key in seen:
            continue
        seen.add(key)
        files.append(path.resolve())
    return files


def build_mapping(
    source_numbers: list[str], pool: dict[str, EmployeeRow], rng: random.Random
) -> dict[str, EmployeeRow]:
    """Построить биективное соответствие исходных табельных и замен."""
    unique_sources: list[str] = sorted(set(source_numbers))
    available_rows: list[EmployeeRow] = list(pool.values())
    rng.shuffle(available_rows)
    if len(unique_sources) > len(available_rows):
        raise RuntimeError(
            f"Недостаточно табельных в пуле замены: нужно {len(unique_sources)}, доступно {len(available_rows)}"
        )
    mapping: dict[str, EmployeeRow] = {}
    for idx, src in enumerate(unique_sources):
        mapping[src] = available_rows[idx]
    return mapping


def replace_in_document(
    payload: Any, mapping: dict[str, EmployeeRow], tb_to_ter: dict[str, str]
) -> tuple[int, int]:
    """Заменить employeeNumber и связанные поля в одном JSON-документе."""
    replaced_total: int = 0
    replaced_emp_nodes: int = 0
    for node in walk_employee_nodes(payload):
        src_raw: Any = node.get("employeeNumber", "")
        src_norm: str = normalize_person_number(src_raw)
        if not src_norm or src_norm not in mapping:
            continue
        target: EmployeeRow = mapping[src_norm]
        node["employeeNumber"] = target.person_number_8
        replaced_total += 1
        replaced_emp_nodes += 1
        if "lastName" in node:
            node["lastName"] = target.surname
        if "firstName" in node:
            node["firstName"] = target.first_name
        if "gosbCode" in node:
            node["gosbCode"] = target.gosb_code
        if "tbCode" in node:
            node["tbCode"] = target.tb_code
        if "terDivisionName" in node:
            node["terDivisionName"] = str(
                tb_to_ter.get(target.tb_code, node.get("terDivisionName", ""))
            )
    return replaced_total, replaced_emp_nodes


def collect_occurrences(
    docs: list[tuple[Path, Any]],
) -> list[ReplacementOccurrence]:
    """Собрать вхождения до замены: файл, ID новости и снимок заменяемых полей."""
    result: list[ReplacementOccurrence] = []
    for path, payload in docs:
        for node, news_id in iter_employee_occurrences(payload):
            source_emp: str = normalize_person_number(node.get("employeeNumber", ""))
            if not source_emp:
                continue
            result.append(
                ReplacementOccurrence(
                    source_emp=source_emp,
                    file_name=path.name,
                    news_id=str(news_id or "").strip(),
                    last_name=str(node.get("lastName", "") or "").strip(),
                    first_name=str(node.get("firstName", "") or "").strip(),
                    gosb_code=str(node.get("gosbCode", "") or "").strip(),
                    tb_code=str(node.get("tbCode", "") or "").strip(),
                    ter_division_name=str(node.get("terDivisionName", "") or "").strip(),
                )
            )
    return result


def multiline_join(values: list[str]) -> str:
    """Склеить уникальные значения через перевод строки (порядок стабильный)."""
    unique: list[str] = []
    seen: set[str] = set()
    for value in values:
        text: str = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        unique.append(text)
    return "\n".join(unique)


def target_after_values(
    target: EmployeeRow, tb_to_ter: dict[str, str]
) -> dict[str, str]:
    """Значения «стало» по заменяемым полям из строки пула."""
    return {
        "employeeNumber": target.person_number_8,
        "lastName": target.surname,
        "firstName": target.first_name,
        "gosbCode": target.gosb_code,
        "tbCode": target.tb_code,
        "terDivisionName": str(tb_to_ter.get(target.tb_code, "")),
    }


def write_excel_report(
    report_path: Path,
    mapping: dict[str, EmployeeRow],
    occurrences: list[ReplacementOccurrence],
    tb_to_ter: dict[str, str],
) -> None:
    """
    Записать Excel: все поля «было», затем все «стало», затем ID новостей, затем файлы.
    Несколько значений/файлов/ID — через перевод строки; заголовок зафиксирован + автофильтр.
    """
    files_by_src: dict[str, list[str]] = defaultdict(list)
    news_by_src: dict[str, list[str]] = defaultdict(list)
    before_by_src: dict[str, dict[str, list[str]]] = defaultdict(
        lambda: {key: [] for key in BEFORE_FIELD_KEYS}
    )
    for hit in occurrences:
        if hit.source_emp not in mapping:
            continue
        files_by_src[hit.source_emp].append(hit.file_name)
        if hit.news_id:
            news_by_src[hit.source_emp].append(hit.news_id)
        before_by_src[hit.source_emp]["employeeNumber"].append(hit.source_emp)
        before_by_src[hit.source_emp]["lastName"].append(hit.last_name)
        before_by_src[hit.source_emp]["firstName"].append(hit.first_name)
        before_by_src[hit.source_emp]["gosbCode"].append(hit.gosb_code)
        before_by_src[hit.source_emp]["tbCode"].append(hit.tb_code)
        before_by_src[hit.source_emp]["terDivisionName"].append(hit.ter_division_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Замены"

    header_font = Font(bold=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False, vertical="center")

    row_idx: int = 2
    for source_emp in sorted(mapping.keys()):
        target: EmployeeRow = mapping[source_emp]
        after: dict[str, str] = target_after_values(target, tb_to_ter)
        before_map: dict[str, list[str]] = before_by_src.get(
            source_emp, {key: [] for key in BEFORE_FIELD_KEYS}
        )
        before_vals: list[str] = [
            multiline_join(before_map.get(key, [])) for key in BEFORE_FIELD_KEYS
        ]
        after_vals: list[str] = [after[key] for key in BEFORE_FIELD_KEYS]
        news_text: str = multiline_join(news_by_src.get(source_emp, []))
        files_text: str = multiline_join(files_by_src.get(source_emp, []))
        values: list[str] = before_vals + after_vals + [news_text, files_text]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align
        line_count: int = 1
        for text in values:
            if text:
                line_count = max(line_count, text.count("\n") + 1)
        ws.row_dimensions[row_idx].height = min(15 * line_count, 150)
        row_idx += 1

    last_row: int = max(1, row_idx - 1)
    last_col: str = get_column_letter(len(EXCEL_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"

    widths: tuple[int, ...] = (
        16,
        16,
        14,
        12,
        10,
        16,
        16,
        16,
        14,
        12,
        10,
        16,
        28,
        28,
    )
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(report_path)


def main() -> None:
    """Точка входа CLI."""
    parser = argparse.ArgumentParser(
        description="Замена employeeNumber в JSON-файлах. "
        "Без аргументов берёт config_emp_replace.json рядом со скриптом."
    )
    parser.add_argument(
        "--config",
        required=False,
        default=None,
        help="Опционально: путь к config_<name>.json (по умолчанию автопоиск рядом со скриптом)",
    )
    args = parser.parse_args()

    try:
        config_path: Path = resolve_config_path(args.config)
    except FileNotFoundError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(2) from exc

    cfg: dict[str, Any] = load_config(config_path)
    base_dir: Path = config_path.parent

    input_dir: Path = (base_dir / str(cfg.get("input_dir", "IN"))).resolve()
    output_dir: Path = (base_dir / str(cfg.get("output_dir", "OUT"))).resolve()
    csv_path: Path = (
        base_dir / str(cfg.get("employee_csv_file", "custom_cib_kksb_dvl.dm_gamification_list_employee.csv"))
    ).resolve()
    output_prefix: str = str(cfg.get("output_prefix", "REPL_EmpID_"))
    excel_report_name: str = str(
        cfg.get("excel_report_file", "REPL_EmpID_mapping.xlsx")
    )
    seed: int = int(cfg.get("random_seed", 20260805))
    tb_to_ter: dict[str, str] = {
        str(k): str(v) for k, v in dict(cfg.get("tb_to_ter_division_name", {})).items()
    }

    print(f"Конфиг: {config_path}")
    try:
        files: list[Path] = collect_input_files(input_dir, cfg)
    except FileNotFoundError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(2) from exc

    if not files:
        print(
            f"[ERROR] Нет входных файлов. Задайте input_files в конфиге "
            f"или положите JSON в {input_dir}",
            file=sys.stderr,
        )
        raise SystemExit(2)

    if not csv_path.is_file():
        print(f"[ERROR] CSV-справочник не найден: {csv_path}", file=sys.stderr)
        raise SystemExit(2)

    output_dir.mkdir(parents=True, exist_ok=True)
    pool: dict[str, EmployeeRow] = load_employee_pool(csv_path, cfg)
    if not pool:
        print("[ERROR] CSV не дал ни одной записи после фильтрации", file=sys.stderr)
        raise SystemExit(2)

    docs: list[tuple[Path, Any]] = []
    all_source_numbers: list[str] = []
    for path in files:
        with path.open("r", encoding="utf-8") as fp:
            payload: Any = json.load(fp)
        docs.append((path, payload))
        for node in walk_employee_nodes(payload):
            normalized: str = normalize_person_number(node.get("employeeNumber", ""))
            if normalized:
                all_source_numbers.append(normalized)

    if not all_source_numbers:
        print("[ERROR] Во входных файлах не найдено ни одного employeeNumber", file=sys.stderr)
        raise SystemExit(2)

    occurrences: list[ReplacementOccurrence] = collect_occurrences(docs)

    rng = random.Random(seed)
    mapping: dict[str, EmployeeRow] = build_mapping(all_source_numbers, pool, rng)

    total_replaced: int = 0
    for path, payload in docs:
        replaced_total, _ = replace_in_document(payload, mapping, tb_to_ter)
        total_replaced += replaced_total
        out_path: Path = output_dir / f"{output_prefix}{path.name}"
        with out_path.open("w", encoding="utf-8") as fp:
            json.dump(payload, fp, ensure_ascii=False, indent=2)
            fp.write("\n")
        print(f"[OK] {path.name} -> {out_path.name} | замен: {replaced_total}")

    report_path: Path = Path(excel_report_name)
    if not report_path.is_absolute():
        report_path = output_dir / report_path
    write_excel_report(report_path, mapping, occurrences, tb_to_ter)
    print(f"[OK] Excel-отчёт: {report_path.name} | строк: {len(mapping)}")

    print(
        f"Готово. Файлов: {len(docs)}, уникальных исходных employeeNumber: {len(set(all_source_numbers))}, "
        f"всего замен: {total_replaced}, размер пула: {len(pool)}"
    )


if __name__ == "__main__":
    main()
